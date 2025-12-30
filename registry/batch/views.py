from datetime import datetime
from difflib import get_close_matches
from io import StringIO

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
from werkzeug.datastructures import Headers
from werkzeug.wrappers import Response

from registry.donor.models import (
    Batch,
    ContactImportLog,
    DonorsOverview,
    Note,
    Record,
)
from registry.extensions import db
from registry.list.models import DonationCenter
from registry.utils import (
    flash_errors,
    get_empty_str_if_none,
    record_as_input_data,
)

from .forms import ContactImportForm, DeleteBatchForm, ImportForm
from .utils import (
    convert_csv_to_text,
    convert_xlsx_to_text,
    process_contact_import_line,
)

blueprint = Blueprint("batch", __name__, static_folder="../static")


@blueprint.get("/import/")
@blueprint.get("/import/<rodne_cislo>")
@login_required
def import_data(rodne_cislo=None):
    import_form = ImportForm()
    if rodne_cislo:
        donation_center_id = request.args.get("donation_center")
        donation_center = DonationCenter.query.get(donation_center_id)
        import_form.donation_center_id.default = donation_center_id
        # Special case for manual imports - empty value in db
        if donation_center_id == "-1":
            donation_center_db_id = None
        else:
            donation_center_db_id = donation_center_id
        import_form.process()
        last_record = (
            Record.query.join(Batch)
            .filter(Record.rodne_cislo == rodne_cislo)
            .filter(Batch.donation_center_id == donation_center_db_id)
            .order_by(Batch.imported_at.desc())
            .first()
        )
        if last_record is not None and (
            donation_center_db_id is None or donation_center.import_increments
        ):
            import_form.input_data.data = record_as_input_data(
                last_record, donation_count="_POČET_", sum_with_last=True
            )
        else:
            # No last record for the given donation center
            # take the last one and start with zero.
            last_record = (
                Record.query.join(Batch)
                .filter(Record.rodne_cislo == rodne_cislo)
                .order_by(Batch.imported_at.desc())
                .first()
            )
            import_form.input_data.data = record_as_input_data(
                last_record, donation_count="_POČET_"
            )

    return render_template("batch/import.html", form=import_form)


@blueprint.post("/import/")
@login_required
def import_data_post():
    import_form = ImportForm(request.form)
    if import_form.validate_on_submit():
        batch = Batch(
            donation_center_id=(
                import_form.donation_center_id.data
                if import_form.donation_center_id.data != "-1"
                else None
            ),
            imported_at=datetime.now(),
        )
        db.session.add(batch)
        db.session.commit()

        for line in import_form.valid_lines_content:
            record = Record.from_list([batch.id] + line.split(";"))
            db.session.add(record)
        db.session.commit()
        # After successfull import, refresh overview table
        DonorsOverview.refresh_overview()
        flash("Import proběhl úspěšně", "success")
        if len(import_form.valid_lines_content) == 1:
            return redirect(url_for("donor.detail", rc=record.rodne_cislo))
        else:
            return redirect(url_for("donor.overview"))
    else:
        flash_errors(import_form)
        return render_template("batch/import.html", form=import_form)


@blueprint.get("/batch_list")
@login_required
def batch_list():
    batches = Batch.query.all()
    delete_batch_form = DeleteBatchForm()
    return render_template(
        "batch/batch_list.html", batches=batches, delete_batch_form=delete_batch_form
    )


@blueprint.post("/delete_batch")
@login_required
def delete_batch():
    delete_batch_form = DeleteBatchForm()
    if delete_batch_form.validate_on_submit():
        records = Record.query.filter(Record.batch_id == delete_batch_form.batch.id)
        for record in records:
            db.session.delete(record)
        db.session.delete(delete_batch_form.batch)
        db.session.commit()
        DonorsOverview.refresh_overview()
        flash("Dávka smazána.", "success")
    else:
        flash("Při odebrání dávky došlo k chybě.", "danger")
    return redirect(url_for("batch.batch_list"))


@blueprint.get("/batch_detail/<id>")
@login_required
def batch_detail(id):
    batch = db.get_or_404(Batch, id)
    records = Record.query.filter(Record.batch_id == batch.id)
    delete_batch_form = DeleteBatchForm()
    return render_template(
        "batch/batch_detail.html",
        batch=batch,
        records=records,
        delete_batch_form=delete_batch_form,
    )


@blueprint.route("/download_batch/<id>", methods=("GET",))
@login_required
def download_batch(id):
    content = StringIO()
    for record in Record.query.filter(Record.batch_id == id):
        content.write(record_as_input_data(record))

    content.seek(0)

    headers = Headers()
    headers.set("Content-Disposition", "attachment", filename="data.txt")

    return Response(content, mimetype="text/plain", headers=headers)


@blueprint.post("/prepare_data_from_trinec")
@login_required
def prepare_data_from_trinec():
    """Process data from Třinec Excel file."""
    if "trinec_file" not in request.files:
        flash("Nebyl vybrán žádný soubor", "danger")
        return redirect(url_for("batch.import_data"))

    file = request.files["trinec_file"]

    if file.filename == "":
        flash("Nebyl vybrán žádný soubor", "danger")
        return redirect(url_for("batch.import_data"))

    try:
        # Load the workbook
        workbook = load_workbook(filename=file, read_only=True)
        sheet = workbook.active

        # Read header row (first row)
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [h.strip() for h in next(rows_iter) if h]

        # Process data rows as dictionaries
        data_rows = []
        for row in rows_iter:
            # Skip empty rows
            if not any(row):
                continue
            # Create dictionary with headers as keys
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            data_rows.append(row_dict)

        workbook.close()

        try:
            # Maps real column names in the Excel to expected names
            column_map = {
                "rodné číslo": get_close_matches("Rodné číslo", headers, n=1)[0],
                "jméno": get_close_matches("Jméno", headers, n=1)[0],
                "příjmení": get_close_matches("Příjmení", headers, n=1)[0],
                "ulice": get_close_matches("TB ulice", headers, n=1)[0],
                "město": get_close_matches("TB město", headers, n=1)[0],
                "psč": get_close_matches("TB psč", headers, n=1)[0],
                "pojišťovna": get_close_matches("Pojišť.", headers, n=1)[0],
                "odběr": get_close_matches("Odběr poř.číslo", headers, n=1)[0],
            }
        except IndexError:
            flash(
                "Při zpracování souboru došlo k chybě: Sloupce se nepodařilo rozpoznat",
                "danger",
            )
            return redirect(url_for("batch.import_data"))

        input_lines = []
        for row in data_rows:
            rodne_cislo = str(row.get(column_map["rodné číslo"]))
            if not rodne_cislo.isnumeric():
                continue
            # Example mapping - adjust these column names to match your Excel file:
            line = ";".join(
                [
                    str(rodne_cislo),
                    str(get_empty_str_if_none(row, column_map["jméno"])),
                    str(get_empty_str_if_none(row, column_map["příjmení"])),
                    str(get_empty_str_if_none(row, column_map["ulice"])),
                    str(get_empty_str_if_none(row, column_map["město"])),
                    str(get_empty_str_if_none(row, column_map["psč"])),
                    str(get_empty_str_if_none(row, column_map["pojišťovna"])),
                    str(get_empty_str_if_none(row, column_map["odběr"])),
                ]
            )
            input_lines.append(line)

        # Create the input data text
        input_data_text = "\n".join(input_lines)

        # Create and pre-populate the ImportForm
        import_form = ImportForm()
        import_form.input_data.data = input_data_text
        import_form.donation_center_id.data = "3"  # Třinec

        flash(
            f"Soubor byl úspěšně načten. Nalezeno {len(input_lines)} řádků."
            f" ({len(data_rows) - len(input_lines)} řádků vynecháno)",
            "success",
        )

        # Render the import template with pre-filled form
        return render_template("batch/import.html", form=import_form)

    except Exception as e:  # noqa: B902
        flash(f"Při zpracování souboru došlo k chybě: {str(e)}", "danger")
        return redirect(url_for("batch.import_data"))


@blueprint.get("/import_contacts/")
@login_required
def import_contacts():
    """Display contact import form."""
    contact_form = ContactImportForm()
    return render_template("batch/import_contacts.html", form=contact_form)


@blueprint.post("/import_contacts_post")
@login_required
def import_contacts_post():
    """Process contact import form submission."""
    contact_form = ContactImportForm(request.form)

    if contact_form.validate_on_submit():
        # Process valid lines
        stats = {
            "total": 0,
            "new_notes": 0,
            "emails_added": 0,
            "phones_added": 0,
            "emails_skipped": 0,
            "phones_skipped": 0,
        }

        for line in contact_form.valid_lines_content:
            data = process_contact_import_line(line)
            rodne_cislo = data["rodne_cislo"]
            email = data["email"]
            phone = data["phone"]

            stats["total"] += 1

            # Get or create note
            note = Note.query.get(rodne_cislo)
            note_is_new = False
            note_updated = False

            if not note:
                note = Note(rodne_cislo=rodne_cislo, note="")
                note_is_new = True
                stats["new_notes"] += 1

            # Check and add email
            if email:
                if email in note.note:
                    stats["emails_skipped"] += 1
                else:
                    if note.note:
                        note.note += "\n" + email
                    else:
                        note.note = email
                    stats["emails_added"] += 1
                    note_updated = True

            # Check and add phone
            if phone:
                if phone in note.note:
                    stats["phones_skipped"] += 1
                else:
                    if note.note:
                        note.note += "\n" + phone
                    else:
                        note.note = phone
                    stats["phones_added"] += 1
                    note_updated = True

            # Save note if updated or new
            if note_is_new or note_updated:
                db.session.add(note)

        db.session.commit()

        # Create audit log entry
        # Determine the input data that was processed
        if contact_form.valid_lines.data or contact_form.invalid_lines.data:
            # Repeated import with fixed errors - use valid lines only
            input_data_logged = contact_form.valid_lines.data
        else:
            # First import - use original input
            input_data_logged = contact_form.input_data.data

        audit_log = ContactImportLog(
            imported_at=datetime.now(),
            imported_by_user_id=current_user.id,
            filename=contact_form.filename.data if contact_form.filename.data else None,
            input_data=input_data_logged,
            processed_lines_count=stats["total"],
            created_notes_count=stats["new_notes"],
            updated_notes_count=stats["total"] - stats["new_notes"],
            emails_added_count=stats["emails_added"],
            phones_added_count=stats["phones_added"],
        )
        db.session.add(audit_log)
        db.session.commit()

        # Flash success message with statistics
        flash(
            f"Import kontaktů proběhl úspěšně. "
            f"Zpracováno: {stats['total']} řádků, "
            f"Nových poznámek: {stats['new_notes']}, "
            f"E-mailů přidáno: {stats['emails_added']} (přeskočeno: {stats['emails_skipped']}), "
            f"Telefonů přidáno: {stats['phones_added']} (přeskočeno: {stats['phones_skipped']})",
            "success",
        )

        return redirect(url_for("donor.overview"))
    else:
        flash_errors(contact_form)
        return render_template("batch/import_contacts.html", form=contact_form)


@blueprint.post("/prepare_contacts_from_file")
@login_required
def prepare_contacts_from_file():
    """Process uploaded file and convert to text format."""
    if "contact_file" not in request.files:
        flash("Nebyl vybrán žádný soubor", "danger")
        return redirect(url_for("batch.import_contacts"))

    file = request.files["contact_file"]

    if file.filename == "":
        flash("Nebyl vybrán žádný soubor", "danger")
        return redirect(url_for("batch.import_contacts"))

    try:
        filename = file.filename.lower()

        # Determine file type and convert
        if filename.endswith(".xlsx"):
            input_data_text = convert_xlsx_to_text(file)
        elif filename.endswith(".csv"):
            input_data_text = convert_csv_to_text(file)
        elif filename.endswith(".txt"):
            # Text file - read directly
            input_data_text = file.read().decode("utf-8")
        else:
            flash(
                "Nepodporovaný formát souboru. Použijte .txt, .csv nebo .xlsx", "danger"
            )
            return redirect(url_for("batch.import_contacts"))

        # Create and pre-populate the form
        contact_form = ContactImportForm()
        contact_form.input_data.data = input_data_text
        contact_form.filename.data = file.filename

        line_count = len(input_data_text.strip().splitlines())
        flash(
            f"Soubor byl úspěšně načten. Nalezeno {line_count} řádků.",
            "success",
        )

        return render_template("batch/import_contacts.html", form=contact_form)

    except Exception as e:  # noqa: B902
        flash(f"Při zpracování souboru došlo k chybě: {str(e)}", "danger")
        return redirect(url_for("batch.import_contacts"))


@blueprint.get("/contact_import_logs")
@login_required
def contact_import_logs():
    """Display list of contact import audit logs."""
    logs = ContactImportLog.query.order_by(ContactImportLog.imported_at.desc()).all()
    return render_template("batch/contact_import_logs.html", logs=logs)
