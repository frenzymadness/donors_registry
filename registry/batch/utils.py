import re

from registry.utils import is_valid_rc


def get_part_of_line(line, delimiter=";"):
    try:
        del_index = line.index(delimiter)
        return line[:del_index], line[del_index + 1 :]
    except ValueError:
        return line, ""


def is_line_valid(line):
    parts = line.split(";")
    if not parts[-1] or parts[-1] == "0":
        # If the last part of the line is empty we can skip the line entirely.
        # Line ending  with semicolon means that the donor has 0 dotations.
        # The same applies for explicitly mentioned 0 donations.
        return None

    if len(parts) == 8 and all(parts):
        return True
    else:
        return False


def repair_two_semicolons(line):
    repaired_line = line
    while ";;" in repaired_line:
        repaired_line = repaired_line.replace(";;", ";")
    return repaired_line


def repair_line_part_by_part(line):
    errors = []

    # In the case when the count of fields is not
    # correct, there is no reason to continue.
    # This function would ignore the additional fields
    # which is not the correct way to fix a line.
    parts_count = len(line.split(";"))
    if parts_count < 8:
        errors.append("nedostatek polí")
        return line, errors
    elif parts_count > 8:
        errors.append("nadbytek polí")
        return line, errors

    rodne_cislo, rest = get_part_of_line(line)
    if not rodne_cislo:
        errors.append("chybí rodné číslo")
    elif not rodne_cislo.isnumeric():
        errors.append("rodné číslo není číselné")
    elif len(rodne_cislo) > 10:
        errors.append("rodné číslo je příliš dlouhé")
    elif len(rodne_cislo) < 9:
        errors.append("rodné číslo je příliš krátké")

    first_name, rest = get_part_of_line(rest)
    if not first_name:
        errors.append("chybí jméno")

    last_name, rest = get_part_of_line(rest)
    if not last_name:
        errors.append("chybí příjmení")

    address, rest = get_part_of_line(rest)
    if not address:
        errors.append("chybí ulice")

    city, rest = get_part_of_line(rest)
    if not city:
        errors.append("chybí město")

    postal_code, rest = get_part_of_line(rest)
    if not postal_code:
        postal_code = "00000"
        errors.append("chybí PSČ, nahrazeno nulami")

    kod_pojistovny, rest = get_part_of_line(rest)
    if not kod_pojistovny:
        kod_pojistovny = "000"
        errors.append("chybí pojišťovna, nahrazena nulami")

    donation_count, rest = get_part_of_line(rest)
    if not donation_count.isnumeric():
        if m := re.match(r"(^\d+)\+(\d+)$", donation_count):
            d1, d2 = m.groups()
            s = int(d1) + int(d2)
            errors.append(f"vstup {donation_count} sečten = {s}")
            donation_count = str(s)
        else:
            errors.append("nevalidní počet odběrů")

    repaired_line = ";".join(
        [
            rodne_cislo,
            first_name,
            last_name,
            address,
            city,
            postal_code,
            kod_pojistovny,
            donation_count,
        ]
    )

    return repaired_line, errors


def validate_import_data(text_input):
    valid_lines = []  # List of valid lines (strings)
    invalid_lines = []  # List of tuples (line, list of comments)
    for line in text_input.splitlines():
        if is_line_valid(line) is None:
            # None means we should skip the line because the donations count
            # is not present at the end of the line
            continue

        if ";;" in line:
            repaired_line = repair_two_semicolons(line)
            if is_line_valid(repaired_line):
                invalid_lines.append(
                    (repaired_line, ["řádek obsahoval dvojici středníků"])
                )
                continue

        repaired_line, errors = repair_line_part_by_part(line)
        if errors or not is_line_valid(line):
            invalid_lines.append((repaired_line, errors))
        else:
            valid_lines.append(line)

    return valid_lines, invalid_lines


def parse_contact_line(line):
    """
    Parse a line to extract rodne cislo, email, and phone number.
    Returns: tuple (rodne_cislo, email, phone, errors)
    """
    from registry.donor.models import DonorsOverview
    from registry.utils import EMAIL_RE, PHONE_RE, RC_RE

    errors = []
    rodne_cislo = None
    email = None
    phone = None

    # Extract rodne cislo (mandatory)
    rc_matches = re.findall(RC_RE, line)
    rc_matches = [rc for rc in rc_matches if is_valid_rc(rc)]
    if not rc_matches:
        errors.append("chybí rodné číslo")
    elif len(rc_matches) > 1:
        errors.append("více než jedno rodné číslo")
    else:
        # Clean rodne cislo (remove slash and spaces)
        rodne_cislo = rc_matches[0].replace("/", "").replace(" ", "")

        # Remove RC from line to avoid ambiguity with phone numbers
        line = line.replace(rodne_cislo, "")

        # Check if donor exists
        donor = DonorsOverview.query.get(rodne_cislo)
        if not donor:
            errors.append("dárce s tímto rodným číslem neexistuje")

    # Extract email (optional)
    email_matches = re.findall(EMAIL_RE, line)
    if len(email_matches) == 1:
        email = email_matches[0]
    elif len(email_matches) > 1:
        errors.append("více než jeden e-mail")

    # Extract phone (optional)
    phone_matches = re.findall(PHONE_RE, line)
    if len(phone_matches) == 1:
        phone = phone_matches[0]
        # Normalize phone number (remove spaces)
        phone = re.sub(r"\s+", "", phone)
    elif len(phone_matches) > 1:
        errors.append("více než jedno telefonní číslo")

    # At least one contact method must be present
    if not email and not phone and not errors:
        errors.append("chybí e-mail nebo telefon")

    return rodne_cislo, email, phone, errors


def validate_contact_import_data(text_input):
    """
    Validate contact import data.
    Returns: (valid_lines, invalid_lines)
    - valid_lines: list of strings (original lines)
    - invalid_lines: list of tuples (line, list of errors)
    """
    valid_lines = []
    invalid_lines = []

    for line in text_input.splitlines():
        # Skip empty lines
        if not line.strip():
            continue

        rodne_cislo, email, phone, errors = parse_contact_line(line)

        if errors:
            invalid_lines.append((line, errors))
        else:
            valid_lines.append(line)

    return valid_lines, invalid_lines


def process_contact_import_line(line):
    """
    Process a validated contact import line and return structured data.
    Returns: dict with keys: rodne_cislo, email, phone
    """
    rodne_cislo, email, phone, errors = parse_contact_line(line)

    return {"rodne_cislo": rodne_cislo, "email": email, "phone": phone}


def convert_xlsx_to_text(file):
    """
    Convert XLSX file to plain text, one row per line.
    Returns: string with all rows joined by newlines
    """
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file, read_only=True)
    sheet = workbook.active

    lines = []
    for row in sheet.iter_rows(values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        # Join all non-None cells with spaces
        row_text = " ".join(str(cell) for cell in row if cell is not None)
        lines.append(row_text)

    workbook.close()
    return "\n".join(lines)


def convert_csv_to_text(file, encoding="utf-8"):
    """
    Convert CSV file to plain text, one row per line.
    Returns: string with all rows joined by newlines
    """
    import csv
    from io import StringIO

    # Read file content
    content = file.read().decode(encoding)

    # Parse CSV
    reader = csv.reader(StringIO(content))
    lines = []
    for row in reader:
        # Join all cells with spaces
        row_text = " ".join(cell for cell in row if cell)
        if row_text.strip():
            lines.append(row_text)

    return "\n".join(lines)
