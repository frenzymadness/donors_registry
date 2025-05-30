import json
from datetime import datetime
from io import BytesIO
from itertools import chain
from tempfile import NamedTemporaryFile

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import login_required
from flask_weasyprint import CSS, HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, collate, extract
from werkzeug.wrappers import Response

from registry.extensions import db
from registry.list.models import DonationCenter, Medals
from registry.utils import (
    donor_as_row,
    flash_errors,
    get_list_of_images,
    send_email_with_award_doc,
)

from .forms import (
    AwardMedalForm,
    DonorsOverrideForm,
    IgnoreDonorForm,
    NoteForm,
    PrintEnvelopeLabelsForm,
    RemoveFromIgnoredForm,
    RemoveMedalForm,
)
from .models import (
    AwardedMedals,
    DonorsOverride,
    DonorsOverview,
    IgnoredDonors,
    Note,
    Record,
)

blueprint = Blueprint("donor", __name__, static_folder="../static")


@blueprint.get("/awarded/")
@login_required
def awarded():
    awarded_medals = AwardedMedals.query.order_by(AwardedMedals.awarded_at.desc()).all()
    years = set()
    for medal in awarded_medals:
        if medal.awarded_at:
            years.add(medal.awarded_at.year)
    years = sorted(years, reverse=True)
    years.append("")  # For the old system
    return render_template(
        "donor/awarded.html",
        years=years,
        column_names=DonorsOverview.frontend_column_names,
        override_column_names=json.dumps(DonorsOverview.basic_fields),
    )


@blueprint.get("/overview/")
@login_required
def overview():
    return render_template(
        "donor/overview.html",
        column_names=DonorsOverview.frontend_column_names,
        override_column_names=json.dumps(DonorsOverview.basic_fields),
    )


@blueprint.get("/overview/data")
@blueprint.get("/overview/data/year/<int:year>/medal/<medal_slug>")
@login_required
def overview_data(year=None, medal_slug=None):
    """JSON end point for JS Datatable"""
    filter_ = True
    params = request.args.to_dict()

    # Block listing only users with awarded medal in the selected year
    if year is not None:
        year = year if year else None
        medal = Medals.query.filter(Medals.slug == medal_slug).first()
        awarded_medals = AwardedMedals.query.filter(
            and_(
                extract("year", AwardedMedals.awarded_at) == year,
                AwardedMedals.medal_id == medal.id,
            ),
        ).all()
        rodna_cisla = [am.rodne_cislo for am in awarded_medals]
        filter_ = DonorsOverview.rodne_cislo.in_(rodna_cisla)
    all_records_count = DonorsOverview.query.filter(filter_).count()

    # WHERE part
    if params["search[value]"]:
        filter_ = and_(
            filter_,
            DonorsOverview.get_filter_for_search(params["search[value]"]),
        )

    # ORDER BY part
    order_by = DonorsOverview.get_order_by_for_column_id(
        int(params["order[0][column]"]), params["order[0][dir]"]
    )

    # LIMIT, OFFSET
    limit = int(params["length"])
    offset = int(params["start"])

    # Final query without limits to see how many records we have after filtering
    # this number is important for pagination
    filtered_records_count = (
        DonorsOverview.query.outerjoin(DonorsOverview.note)
        .filter(filter_)
        .order_by(*order_by)
        .count()
    )
    # Final query
    overview = (
        DonorsOverview.query.outerjoin(DonorsOverview.note)
        .filter(filter_)
        .order_by(*order_by)
        .limit(limit)
        .offset(offset)
        .all()
    )

    # Data processing
    final_list = []
    for donor in overview:
        final_list.append(donor.dict_for_frontend())
    return jsonify(
        {
            "draw": int(params["draw"]),
            "data": final_list,
            "recordsTotal": all_records_count,
            "recordsFiltered": filtered_records_count,
        }
    )


@blueprint.get("/detail/<rc>")
@login_required
def detail(rc):
    remove_medal_form = RemoveMedalForm()
    award_medal_form = AwardMedalForm()
    award_medal_form.add_one_rodne_cislo(rc)
    overview = db.session.get(DonorsOverview, rc)
    if not overview:
        if db.session.get(IgnoredDonors, rc):
            flash("Dárce je ignorován a proto není jeho detail k dispozici.", "danger")
            return redirect(url_for("donor.show_ignored"))
        return abort(404)
    records = Record.query.filter(Record.rodne_cislo == rc).all()
    donation_centers = DonationCenter.query.all()
    awarded_medals = AwardedMedals.query.filter(AwardedMedals.rodne_cislo == rc).all()
    awarded_medals = {medal.medal.id: medal for medal in awarded_medals}
    all_medals = Medals.query.all()
    note_form = NoteForm()
    if overview.note:
        note_form.note.data = overview.note.note
        emails = overview.note.get_emails_from_note()
    else:
        emails = None
    donors_override_form = DonorsOverrideForm()
    donors_override_form.init_fields(rc)

    return render_template(
        "donor/detail.html",
        overview=overview,
        donation_centers=donation_centers,
        records=records,
        awarded_medals=awarded_medals,
        all_medals=all_medals,
        remove_medal_form=remove_medal_form,
        award_medal_form=award_medal_form,
        note_form=note_form,
        emails=emails,
        donors_override_form=donors_override_form,
    )


@blueprint.get("/detail/<rc>/award_document/<medal_slug>/")
@login_required
def render_award_document(rc, medal_slug):
    donor = db.get_or_404(DonorsOverview, rc)
    medal = Medals.query.filter_by(slug=medal_slug).first_or_404()
    awarded_medal = AwardedMedals.query.filter(
        AwardedMedals.rodne_cislo == donor.rodne_cislo,
        AwardedMedals.medal_id == medal.id,
    ).first()
    if "today" in request.args:
        # Today date requested
        awarded_at = datetime.now()
    elif awarded_medal and awarded_medal.awarded_at:
        # Medal is awarded and it's not in the old system.
        # If it is, we don't know the date.
        awarded_at = awarded_medal.awarded_at
    else:
        # Unknown date (medal awarded in old system)
        awarded_at = datetime.now()

    return render_template(
        "donor/award_document.html",
        donors=(donor,),
        medal=medal,
        awarded_at=awarded_at.strftime("%-d. %-m. %Y"),
        stamps=get_list_of_images("stamps"),
        signatures=get_list_of_images("signatures"),
    )


@blueprint.get("/detail/<rc>/email_award_document/<medal_slug>")
@login_required
def email_award_document(rc, medal_slug):
    note = Note.query.get(rc)
    emails = None
    if note:
        emails = note.get_emails_from_note()

    if not emails:
        flash("Dárce nemá v poznámce žádný e-mail.", "danger")
        return redirect(url_for("donor.detail", rc=rc))

    donor = db.get_or_404(DonorsOverview, rc)
    medal = Medals.query.filter_by(slug=medal_slug).first_or_404()
    awarded_medal = AwardedMedals.query.filter(
        AwardedMedals.rodne_cislo == donor.rodne_cislo,
        AwardedMedals.medal_id == medal.id,
    ).first()

    if awarded_medal and awarded_medal.awarded_at:
        # Medal is awarded and it's not in the old system.
        # If it is, we don't know the date.
        awarded_at = awarded_medal.awarded_at
    else:
        # Unknown date (medal awarded in old system)
        awarded_at = datetime.now()

    award_document_html = render_template(
        "donor/award_document.html",
        donors=(donor,),
        medal=medal,
        awarded_at=awarded_at.strftime("%-d. %-m. %Y"),
        stamps=get_list_of_images("stamps"),
        signatures=get_list_of_images("signatures"),
    )

    css = CSS(url=url_for("static", filename="css/award_document.css", _external=True))
    pdf_content = HTML(string=award_document_html).write_pdf(stylesheets=[css])

    send_email_with_award_doc(
        to=emails, award_doc_content=pdf_content, medal=medal, config=current_app.config
    )

    flash("E-mail odeslán.", "success")
    return redirect(url_for("donor.detail", rc=rc))


@blueprint.get("/detail/<rc>/confirmation_document/")
@login_required
def render_confirmation_document(rc):
    donor = db.get_or_404(DonorsOverview, rc)
    awarded_medals_list = []
    awarded_medals = (
        AwardedMedals.query.filter(AwardedMedals.rodne_cislo == rc)
        .order_by(AwardedMedals.medal_id.asc())
        .all()
    )
    for awarded_medal in awarded_medals:
        awarded_medals_list.append(awarded_medal.medal.title)

    return render_template(
        "donor/confirmation_document.html",
        donor=donor,
        date=datetime.now().strftime("%-d. %-m. %Y"),
        awarded_medals=", ".join(awarded_medals_list),
        stamps=get_list_of_images("stamps"),
        signatures=get_list_of_images("signatures"),
    )


@blueprint.get("/award_prep/documents/<medal_slug>/")
@login_required
def render_award_documents_for_award_prep(medal_slug):
    medal = Medals.query.filter(Medals.slug == medal_slug).first_or_404()
    medal_kr3 = Medals.query.filter(Medals.slug == "kr3").first_or_404()
    donors = (
        DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal_slug).is_(False),
            )
        )
        .order_by(collate(DonorsOverview.last_name, "czech").asc())
        .all()
    )

    # Show date of the award only for lower three medals
    awarded_at = datetime.now().strftime("%-d. %-m. %Y") if medal < medal_kr3 else ""

    return render_template(
        "donor/award_document.html",
        donors=donors,
        medal=medal,
        awarded_at=awarded_at,
        stamps=get_list_of_images("stamps"),
        signatures=get_list_of_images("signatures"),
    )


@blueprint.post("/award_prep/envelope_labels")
@login_required
def render_envelope_labels():
    print_envelope_labels_form = PrintEnvelopeLabelsForm()
    if print_envelope_labels_form.validate_on_submit():
        medal = print_envelope_labels_form.medal
        donors = (
            DonorsOverview.query.filter(
                and_(
                    DonorsOverview.donation_count_total >= medal.minimum_donations,
                    getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
                )
            )
            .order_by(collate(DonorsOverview.last_name, "czech").asc())
            .all()
        )

        # To skip already used labels, we prepend some
        # empty donors to the list.
        empty_donor = {
            k: "" for k in ("first_name", "last_name", "address", "city", "postal_code")
        }
        all_donors = chain([empty_donor] * print_envelope_labels_form.skip.data, donors)
        pages = []

        for index, donor in enumerate(all_donors):
            if index % 16 == 0:
                pages.append([])
            pages[-1].append(donor)

        return render_template(
            "donor/envelope_labels.html",
            pages=pages,
        )

    return redirect(request.referrer)


@blueprint.post("/award_prep/envelope")
@login_required
def render_envelope():
    print_envelope_labels_form = PrintEnvelopeLabelsForm()
    if print_envelope_labels_form.validate_on_submit():
        medal = print_envelope_labels_form.medal
        donors = (
            DonorsOverview.query.filter(
                and_(
                    DonorsOverview.donation_count_total >= medal.minimum_donations,
                    getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
                )
            )
            .order_by(collate(DonorsOverview.last_name, "czech").asc())
            .all()
        )

        return render_template(
            "donor/envelope_DL.html",
            donors=donors,
        )

    return redirect(request.referrer)


@blueprint.post("/remove_medal")
@login_required
def remove_medal():
    remove_medal_form = RemoveMedalForm()
    if remove_medal_form.validate_on_submit():
        db.session.delete(remove_medal_form.awarded_medal)
        db.session.commit()
        DonorsOverview.refresh_overview(rodne_cislo=remove_medal_form.rodne_cislo.data)
        flash("Medaile byla úspěšně odebrána.", "success")
    else:
        flash("Při odebrání medaile došlo k chybě.", "danger")
    return redirect(url_for("donor.detail", rc=remove_medal_form.rodne_cislo.data))


@blueprint.get("/award_prep/<medal_slug>")
@login_required
def award_prep(medal_slug):
    medal = Medals.query.filter(Medals.slug == medal_slug).first_or_404()
    donors = DonorsOverview.query.filter(
        and_(
            DonorsOverview.donation_count_total >= medal.minimum_donations,
            getattr(DonorsOverview, "awarded_medal_" + medal_slug).is_(False),
        )
    ).all()
    award_medal_form = AwardMedalForm()
    award_medal_form.add_checkboxes([d.rodne_cislo for d in donors])
    print_envelope_labels_form = PrintEnvelopeLabelsForm()
    return render_template(
        "donor/award_prep.html",
        medal=medal,
        donors=donors,
        award_medal_form=award_medal_form,
        override_column_names=json.dumps(DonorsOverview.basic_fields),
        print_envelope_labels_form=print_envelope_labels_form,
    )


@blueprint.get("/award_prep_download_table/<medal_slug>")
@login_required
def award_prep_download_table(medal_slug):
    medal = Medals.query.filter(Medals.slug == medal_slug).first_or_404()
    donors = DonorsOverview.query.filter(
        and_(
            DonorsOverview.donation_count_total >= medal.minimum_donations,
            getattr(DonorsOverview, "awarded_medal_" + medal_slug).is_(False),
        )
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "roztridit"
    donation_centers = DonationCenter.query.order_by(DonationCenter.slug.desc()).all()
    for dc in donation_centers:
        wb.create_sheet(dc.title)

    for sheetname in wb.sheetnames:
        wb[sheetname].append(
            [
                "Jméno",
                "Příjmení",
                "Datum narození",
                "Adresa",
                "Město",
                "PSČ",
                "Pojišťovna",
                "Odběrná místa",
            ]
        )

    for donor in donors:
        row = donor_as_row(donor)
        dcs = row[-1]
        row[-1] = ", ".join(dcs)
        if len(dcs) == 1:
            wb[dcs[0]].append(row)
        else:
            wb["roztridit"].append(row)

    for sheetname in wb.sheetnames:
        column_letters = tuple(
            get_column_letter(col_number + 1)
            for col_number in range(wb[sheetname].max_column)
        )
        for column_letter in column_letters:
            wb[sheetname].column_dimensions[column_letter].bestFit = True

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        content = BytesIO(tmp.read())

    return Response(
        content,
        headers={
            "Content-Disposition": f"attachment; filename=darci_k_oceneni_{medal.slug}.xlsx",  # noqa
            "Content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # noqa
        },
    )


@blueprint.post("/award_medal")
@login_required
def award_medal():
    award_medal_form = AwardMedalForm()
    award_medal_form.rodna_cisla = request.form.getlist("rodne_cislo")
    if award_medal_form.validate_on_submit():
        for rodne_cislo in award_medal_form.rodna_cisla:
            am = AwardedMedals(
                rodne_cislo=rodne_cislo,
                medal_id=award_medal_form.medal.id,
                awarded_at=datetime.now(),
            )
            db.session.add(am)
            db.session.commit()
            DonorsOverview.refresh_overview(rodne_cislo=rodne_cislo)

        if len(award_medal_form.rodna_cisla) == 1:
            flash("Medaile udělena.", "success")
        else:
            flash("Medaile uděleny.", "success")
    return redirect(request.referrer)


@blueprint.post("/note/save")
@login_required
def save_note():
    note_form = NoteForm()
    note = db.session.get(Note, note_form.rodne_cislo.data)
    if note:
        note.note = note_form.note.data
    else:
        note = Note(rodne_cislo=note_form.rodne_cislo.data, note=note_form.note.data)
    db.session.add(note)
    db.session.commit()
    flash("Poznámka uložena.", "success")
    return redirect(url_for("donor.detail", rc=note_form.rodne_cislo.data))


@blueprint.get("/donor/ignore")
@login_required
def show_ignored():
    ignored = IgnoredDonors.query.all()
    return render_template(
        "donor/ignore_donor.html",
        ignore_form=IgnoreDonorForm(),
        ignored=ignored,
        unignore_form=RemoveFromIgnoredForm(),
    )


@blueprint.post("/donor/ignore/add")
@login_required
def ignore_donor():
    ignore_form = IgnoreDonorForm()
    if ignore_form.validate_on_submit():
        if not db.session.get(IgnoredDonors, ignore_form.rodne_cislo.data):
            ignored = IgnoredDonors(
                rodne_cislo=ignore_form.rodne_cislo.data,
                reason=ignore_form.reason.data,
                ignored_since=datetime.now(),
            )
            DonorsOverview.query.filter(
                DonorsOverview.rodne_cislo == ignore_form.rodne_cislo.data
            ).delete()
            db.session.add(ignored)
            db.session.commit()
            flash("Dárce ignorován.", "success")
        else:
            flash("Dárce již je v seznamu ignorovaných", "danger")
    else:
        flash("Při přidávání do ignorovaných došlo k chybě", "danger")
    return redirect(url_for("donor.show_ignored"))


@blueprint.post("/donor/ignore/remove")
@login_required
def unignore_donor():
    unignore_form = RemoveFromIgnoredForm()
    if unignore_form.validate_on_submit():
        db.session.delete(unignore_form.ignored_donor)
        db.session.commit()
        DonorsOverview.refresh_overview(rodne_cislo=unignore_form.rodne_cislo.data)
        flash("Dárce již není ignorován.", "success")
    else:
        flash("Při odebírání ze seznamu ignorovaných dárců došlo k chybě", "danger")
    return redirect(url_for("donor.show_ignored"))


@blueprint.post("/override/")
@login_required
def save_override():
    form = DonorsOverrideForm()
    delete = "delete_btn" in request.form

    if form.validate_on_submit():
        if not delete:
            # Save the override
            override = DonorsOverride(**form.get_field_data())
            db.session.merge(override)
            db.session.commit()

            DonorsOverview.refresh_overview()
            flash("Výjimka uložena", "success")
        else:
            # Delete the override
            override = db.session.get(DonorsOverride, form.rodne_cislo.data)
            if override is not None:
                db.session.delete(override)
                db.session.commit()

                DonorsOverview.refresh_overview()
                flash("Výjimka smazána", "success")
            else:
                flash("Není co mazat", "warning")
    else:
        flash_errors(form)

    return redirect(url_for("donor.detail", rc=form.rodne_cislo.data))


@blueprint.get("/override/all")
@login_required
def get_overrides():
    overrides_dict = {}

    for override in DonorsOverride.query.all():
        overrides_dict[override.rodne_cislo] = override.to_dict()

    return jsonify(overrides_dict)
