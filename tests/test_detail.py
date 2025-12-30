import re
from datetime import datetime, timedelta
from math import ceil
from tempfile import NamedTemporaryFile

import pytest
from flask import url_for
from openpyxl import load_workbook
from sqlalchemy import and_

from registry.donor.models import (
    AwardedMedals,
    Batch,
    DonationCenter,
    DonorsOverview,
    Note,
    Record,
)
from registry.extensions import db
from registry.list.models import Medals

from .fixtures import sample_of_rc, skip_if_ignored
from .helpers import login


class TestDetail:
    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(50))
    def test_detail(self, user, testapp, rodne_cislo):
        """Just a simple test that the detail page loads for some random donors"""
        skip_if_ignored(rodne_cislo)
        login(user, testapp)
        res = testapp.get(url_for("donor.detail", rc=rodne_cislo))
        assert res.status_code == 200
        assert "<td></td>" not in res
        # Check that the sum of the donations is eqal to the total count
        donations_list = re.search(r">Po캜ty darov치n칤</h[1-6]>.*?</b>", res.text, re.S)
        numbers = re.findall(r"(\d+)[\n<]{1}", donations_list.group())
        numbers = list(map(int, numbers))
        assert sum(numbers[:-1]) == numbers[-1]

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(5))
    def test_save_update_note(self, user, testapp, rodne_cislo):
        skip_if_ignored(rodne_cislo)
        # Make sure no note exists for this RC
        if note := Note.query.get(rodne_cislo) is not None:
            db.session.delete(note)
            db.session.commit()
        existing_notes = Note.query.count()
        login(user, testapp)
        res = testapp.get(url_for("donor.detail", rc=rodne_cislo))
        # New note
        form = res.forms["noteForm"]
        assert form.fields["note"][0].value == ""
        form.fields["note"][0].value = "Lorem ipsum"
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        assert "Lorem ipsum</textarea>" in res.text
        assert Note.query.count() == existing_notes + 1
        # Update existing
        form = res.forms["noteForm"]
        assert form.fields["note"][0].value == "Lorem ipsum"
        form.fields["note"][0].value += " dolor sit amet,"
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        assert "Lorem ipsum dolor sit amet,</textarea>" in res.text
        assert Note.query.count() == existing_notes + 1

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(5))
    def test_emails_in_notes(self, user, testapp, rodne_cislo):
        skip_if_ignored(rodne_cislo)
        login(user, testapp)
        res = testapp.get(url_for("donor.detail", rc=rodne_cislo))
        # No e-mail in the note
        assert "E-mail:" not in res.text
        assert '<a href="mailto:' not in res.text
        # Add first e-mail
        form = res.forms["noteForm"]
        form.fields["note"][0].value += "\nfoo@example.com"
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        assert "E-mail:" in res.text
        assert '<a href="mailto:foo@example.com">foo@example.com</a>' in res.text
        # Add second e-mail
        form = res.forms["noteForm"]
        form.fields["note"][0].value += "\nbar@example.com"
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        assert "E-mail:" in res.text
        assert '<a href="mailto:foo@example.com">foo@example.com</a>' in res.text
        assert '<a href="mailto:bar@example.com">bar@example.com</a>' in res.text

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(10))
    def test_manual_import_prepare(self, user, testapp, rodne_cislo):
        skip_if_ignored(rodne_cislo)
        login(user, testapp)

        donation_centers = DonationCenter.query.all()

        for dc in donation_centers + ["manual"]:
            dc_db_id = dc.id if dc != "manual" else None
            dc_id = dc.id if dc != "manual" else -1
            import_increments = dc.import_increments if dc != "manual" else True
            last_record = (
                Record.query.join(Batch)
                .filter(Record.rodne_cislo == rodne_cislo)
                .filter(Batch.donation_center_id == dc_db_id)
                .order_by(Batch.imported_at.desc())
                .first()
            )
            if last_record is None or not import_increments:
                last_record = (
                    Record.query.join(Batch)
                    .filter(Record.rodne_cislo == rodne_cislo)
                    .order_by(Batch.imported_at.desc())
                    .first()
                )
                expected_eol = ";_PO캛ET_"
            else:
                expected_eol = f";{last_record.donation_count}+_PO캛ET_"

            detail = testapp.get(url_for("donor.detail", rc=rodne_cislo))
            manual_import_form = detail.forms["manualImportForm"]
            manual_import_form.fields["donation_center"][0].select(dc_id)
            import_page = manual_import_form.submit()
            import_form = import_page.forms[0]
            assert import_form.fields["donation_center_id"][0].value == str(dc_id)
            input_data = import_form.fields["input_data"][0].value
            for field in (
                "rodne_cislo",
                "first_name",
                "last_name",
                "address",
                "city",
                "postal_code",
                "kod_pojistovny",
            ):
                assert getattr(last_record, field) + ";" in input_data
            assert expected_eol in input_data


class TestAwardDocument:
    @pytest.mark.parametrize("rodne_cislo", ("391105000", "9701037137", "151008110"))
    def test_award_doc_for_man(self, user, testapp, rodne_cislo):
        overview = db.session.get(DonorsOverview, rodne_cislo)
        medals = Medals.query.all()
        login(user, testapp)
        for medal in medals:
            doc = testapp.get(
                url_for(
                    "donor.render_award_document", rc=rodne_cislo, medal_slug=medal.slug
                )
            )

            assert f"{rodne_cislo[:6]}/{rodne_cislo[6:]}" in doc
            assert overview.first_name in doc
            assert overview.last_name in doc
            assert "pracovn칤k" in doc
            assert "jeho" in doc
            assert "p." in doc
            assert medal.title not in doc
            assert medal.title_acc in doc
            assert medal.title_instr in doc
            assert re.search(r"<img src=\"/static/stamps/.*\.png\"", doc.text)

    @pytest.mark.parametrize("rodne_cislo", ("0457098862", "0552277759", "0160031652"))
    def test_award_doc_for_woman(self, user, testapp, rodne_cislo):
        overview = db.session.get(DonorsOverview, rodne_cislo)
        medals = Medals.query.all()
        login(user, testapp)
        for medal in medals:
            doc = testapp.get(
                url_for(
                    "donor.render_award_document", rc=rodne_cislo, medal_slug=medal.slug
                )
            )

            assert f"{rodne_cislo[:6]}/{rodne_cislo[6:]}" in doc
            assert overview.first_name in doc
            assert overview.last_name in doc
            assert "pracovnice" in doc
            assert "jej칤" in doc
            assert "p칤" in doc
            assert medal.title not in doc
            assert medal.title_acc in doc
            assert medal.title_instr in doc
            assert re.search(r"<img src=\"/static/stamps/.*\.png\"", doc.text)

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(5))
    def test_award_doc_dates(self, user, testapp, db, rodne_cislo):
        today = datetime.now().strftime("%-d. %-m. %Y")
        login(user, testapp)

        # Test medal from the old system, where the date is unknown
        awarded_at = None
        am = AwardedMedals.query.filter(
            AwardedMedals.rodne_cislo == rodne_cislo, AwardedMedals.medal_id == 1
        ).first()
        if am:
            am.awarded_at = awarded_at
        else:
            am = AwardedMedals(
                rodne_cislo=rodne_cislo, medal_id=1, awarded_at=awarded_at
            )
        db.session.add(am)
        db.session.commit()

        doc = testapp.get(
            url_for("donor.render_award_document", rc=rodne_cislo, medal_slug="br")
        )

        assert f"Ve Fr칳dku-M칤stku, dne {today}" in doc

        # Test medal from the new system, where the date is known
        awarded_at = datetime(1989, 11, 17, 12, 23, 12)
        am = AwardedMedals.query.filter(
            AwardedMedals.rodne_cislo == rodne_cislo, AwardedMedals.medal_id == 2
        ).first()
        if am:
            am.awarded_at = awarded_at
        else:
            am = AwardedMedals(
                rodne_cislo=rodne_cislo, medal_id=2, awarded_at=awarded_at
            )
        db.session.add(am)
        db.session.commit()

        doc = testapp.get(
            url_for("donor.render_award_document", rc=rodne_cislo, medal_slug="st")
        )

        assert "Ve Fr칳dku-M칤stku, dne 17. 11. 1989" in doc

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(5))
    def test_award_doc_today(self, user, testapp, db, rodne_cislo):
        """Test that if today is set in request arguments,
        document always content today's date"""
        today = datetime.now().strftime("%-d. %-m. %Y")
        login(user, testapp)
        medals = Medals.query.all()
        for medal in medals:
            doc = testapp.get(
                url_for(
                    "donor.render_award_document",
                    rc=rodne_cislo,
                    medal_slug=medal.slug,
                    today=1,
                )
            )

            assert f"Ve Fr칳dku-M칤stku, dne {today}" in doc

    @pytest.mark.parametrize("medal_id", range(1, 8))
    def test_award_prep_documents(self, user, testapp, medal_id):
        medal = db.session.get(Medals, medal_id)
        medal_kr3 = Medals.query.filter(Medals.slug == "kr3").first_or_404()
        today = datetime.now().strftime("%-d. %-m. %Y") if medal < medal_kr3 else ""
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        rows = page.text.count("<tr") - 1  # Minus 1 for table header
        documents = page.click(description="Potvrzen칤 k medail칤m pro v코echny")

        assert rows == documents.text.count('<div class="page">')
        assert rows == documents.text.count(f"Ve Fr칳dku-M칤stku, dne {today}")

    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(3))
    def test_email_award_document_no_email(self, user, testapp, rodne_cislo):
        skip_if_ignored(rodne_cislo)
        login(user, testapp)
        res = testapp.get(url_for("donor.detail", rc=rodne_cislo))
        # Make sure the note is empty
        form = res.forms["noteForm"]
        form.fields["note"][0].value = ""
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        # There should be no buttons to send e-mails now
        assert "/email_award_document/" not in res
        assert "游닎" not in res
        # If called somehow anyway, proper error is displayed
        res = testapp.get(
            url_for("donor.email_award_document", rc=rodne_cislo, medal_slug="st")
        ).follow()
        assert "D치rce nem치 v pozn치mce 쮂멳n칳 e-mail." in res

    @pytest.mark.parametrize(
        "note,expected_to",
        [
            ("foo@bar.baz", "foo@bar.baz"),
            ("foo@bar.baz\nboss@example.com", "foo@bar.baz, boss@example.com"),
            ("Some random text foo@bar.baz", "foo@bar.baz"),
            ("foo@bar.baz and some random note", "foo@bar.baz"),
            (
                "heading email@example.com heading 2 email2@example.com",
                "email@example.com, email2@example.com",
            ),
        ],
    )
    @pytest.mark.parametrize("medal_id", range(1, 8))
    @pytest.mark.parametrize("rodne_cislo", sample_of_rc(2))
    def test_email_award_document(
        self, user, testapp, note, expected_to, medal_id, rodne_cislo, mock_smtp
    ):
        skip_if_ignored(rodne_cislo)
        medal = db.session.get(Medals, medal_id)
        # Make sure at least some medals are already awarded
        awarded_medals = (
            AwardedMedals.query.filter(AwardedMedals.rodne_cislo == rodne_cislo)
            .order_by(AwardedMedals.medal_id.asc())
            .all()
        )
        if len(awarded_medals) == 0 and medal_id % 2 == 0:
            am = AwardedMedals(
                rodne_cislo=rodne_cislo,
                medal_id=medal_id,
                awarded_at=datetime.now() - timedelta(weeks=1),
            )
            db.session.add(am)
            db.session.commit()
            DonorsOverview.refresh_overview()
        login(user, testapp)
        res = testapp.get(url_for("donor.detail", rc=rodne_cislo))
        # Make sure there is one e-mail in the note
        form = res.forms["noteForm"]
        form.fields["note"][0].value = note
        res = form.submit().follow()
        assert res.status_code == 200
        assert "Pozn치mka ulo쬰na." in res
        # There should be buttons to send e-mails present
        assert "/email_award_document/" in res
        assert "游닎" in res
        # Click on the button
        res = res.click(
            description="游닎", href=f"/email_award_document/{medal.slug}"
        ).follow()
        assert "E-mail odesl치n." in res
        # Assert proper mail has been sent
        ctx_mngr = mock_smtp.__enter__.return_value
        ctx_mngr.login.assert_called_once()
        ctx_mngr.send_message.assert_called_once()
        message = ctx_mngr.send_message.call_args[0][0]
        assert message["to"] == expected_to
        assert message["cc"] == "foo@example.com"
        assert message["subject"] == "Ocen캩n칤 za darov치n칤 krve a krevn칤ch slo쬰k"
        for part in message.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            if content_type == "text/plain" and "attachment" not in content_disposition:
                charset = part.get_content_charset() or "utf-8"
                content = part.get_payload(decode=True).decode(
                    charset, errors="replace"
                )
                if medal.slug in ("br", "st"):
                    assert "si pros칤m vyzvedn캩te" in content
                elif medal.slug in ("zl", "kr3"):
                    assert "se uskute캜n칤 na podzim v T콏inci a Fr칳dku-M칤stku" in content
                elif medal.slug in ("kr2", "kr1", "plk"):
                    assert "Pozv치nku na slavnostn칤 oce켿ov치n칤 obdr쮂셦e" in content


class TestConfirmationdDocument:
    @pytest.mark.parametrize("rodne_cislo", ("391105000", "9701037137", "151008110"))
    def test_confirmation_doc_for_man(self, user, testapp, rodne_cislo):
        overview = db.session.get(DonorsOverview, rodne_cislo)
        login(user, testapp)
        doc = testapp.get(url_for("donor.render_confirmation_document", rc=rodne_cislo))

        assert f"{rodne_cislo[:6]}/{rodne_cislo[6:]}" in doc
        assert f"p. {overview.first_name} {overview.last_name}" in doc
        assert f" {overview.donation_count_total} bezp콏칤sp캩vkov칳ch odb캩r콢" in doc
        assert re.search(r"<img src=\"/static/stamps/.*\.png\"", doc.text)
        assert re.search(r"<img src=\"/static/signatures/.*\.png\"", doc.text)

        awarded_medals = (
            AwardedMedals.query.filter(AwardedMedals.rodne_cislo == rodne_cislo)
            .order_by(AwardedMedals.medal_id.asc())
            .all()
        )
        if len(awarded_medals):
            assert "D치rce je dr쬴telem" in doc
        else:
            assert "D치rce je dr쬴telem" not in doc

        for awarded_medal in awarded_medals:
            assert awarded_medal.medal.title in doc

    @pytest.mark.parametrize("rodne_cislo", ("0457098862", "0552277759", "0160031652"))
    def test_confirmation_doc_for_woman(self, user, testapp, rodne_cislo):
        overview = db.session.get(DonorsOverview, rodne_cislo)
        login(user, testapp)
        doc = testapp.get(url_for("donor.render_confirmation_document", rc=rodne_cislo))

        assert f"{rodne_cislo[:6]}/{rodne_cislo[6:]}" in doc
        assert f"p칤 {overview.first_name} {overview.last_name}" in doc
        assert f" {overview.donation_count_total} bezp콏칤sp캩vkov칳ch odb캩r콢" in doc
        assert re.search(r"<img src=\"/static/stamps/.*\.png\"", doc.text)
        assert re.search(r"<img src=\"/static/signatures/.*\.png\"", doc.text)

        awarded_medals = (
            AwardedMedals.query.filter(AwardedMedals.rodne_cislo == rodne_cislo)
            .order_by(AwardedMedals.medal_id.asc())
            .all()
        )
        if len(awarded_medals):
            assert "D치rkyn캩 je dr쬴telkou" in doc
        else:
            assert "D치rkyn캩 je dr쬴telkou" not in doc

        for awarded_medal in awarded_medals:
            assert awarded_medal.medal.title in doc


class TestEnvelopeLabels:
    @pytest.mark.parametrize("medal_id", range(1, 8))
    def test_envelope_labels(self, user, testapp, medal_id):
        medal = db.session.get(Medals, medal_id)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        labels = page.forms["printEnvelopeLabelsForm"].submit()
        eligible_donors = DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
            )
        ).count()

        pages_count = labels.text.count('<div class="page">')
        assert ceil(eligible_donors / 16) == pages_count

        labels_count = labels.text.count('<div class="label">')
        assert labels_count == eligible_donors

    def test_envelope_labels_detail(self, user, testapp):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        labels = page.forms["printEnvelopeLabelsForm"].submit()
        eligible_donors = DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
            )
        ).all()

        for donor in eligible_donors:
            assert f"<p>{donor.first_name} {donor.last_name}</p>" in labels.text
            assert f"<p>{donor.address}</p>" in labels.text
            assert f"<p>{donor.postal_code} {donor.city}</p>" in labels.text

    @pytest.mark.parametrize("skip", (1, 3, 9, 13))
    @pytest.mark.parametrize("medal_id", range(1, 8))
    def test_envelope_labels_skip(self, user, testapp, medal_id, skip):
        medal = db.session.get(Medals, medal_id)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        page.forms["printEnvelopeLabelsForm"].fields["skip"][0].value = skip
        labels = page.forms["printEnvelopeLabelsForm"].submit()
        eligible_donors = DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
            )
        ).count()

        labels_count = labels.text.count('<div class="label">')
        assert labels_count == eligible_donors + skip

        pages_count = labels.text.count('<div class="page">')
        assert ceil((eligible_donors + skip) / 16) == pages_count

        empty_p = labels.text.count("<p></p>")  # for address
        space_p = labels.text.count("<p> </p>")  # for name and city

        assert empty_p == skip
        assert space_p == skip * 2

    @pytest.mark.parametrize("skip", (-1, -33, 99, 16))
    def test_envelope_labels_invalid_skip(self, user, testapp, skip):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        page.forms["printEnvelopeLabelsForm"].fields["skip"][0].value = skip
        labels = page.forms["printEnvelopeLabelsForm"].submit().follow()

        assert "Vynechat lze 0 a 15 코t칤tk콢." in labels.text

    def test_envelope_labels_empty_skip(self, user, testapp):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        page.forms["printEnvelopeLabelsForm"].fields["skip"][0].value = ""
        labels = page.forms["printEnvelopeLabelsForm"].submit()
        eligible_donors = DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
            )
        ).count()

        labels_count = labels.text.count('<div class="label">')
        assert labels_count == eligible_donors

        pages_count = labels.text.count('<div class="page">')
        assert ceil(eligible_donors / 16) == pages_count

        empty_p = labels.text.count("<p></p>")  # for address
        space_p = labels.text.count("<p> </p>")  # for name and city

        assert empty_p == 0
        assert space_p == 0

    @pytest.mark.parametrize("medal_id", (-1, -33, 99, 16))
    def test_envelope_labels_invalid_medal(self, user, testapp, medal_id):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        page.forms["printEnvelopeLabelsForm"].fields["medal_id"][0].value = medal_id
        labels = page.forms["printEnvelopeLabelsForm"].submit().follow()

        assert "Odesl치na nevalidn칤 data." in labels.text

    def test_envelope_detail(self, user, testapp):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        labels = page.forms["printEnvelopeForm"].submit()
        eligible_donors = DonorsOverview.query.filter(
            and_(
                DonorsOverview.donation_count_total >= medal.minimum_donations,
                getattr(DonorsOverview, "awarded_medal_" + medal.slug).is_(False),
            )
        ).all()

        for donor in eligible_donors:
            assert f"{donor.first_name} {donor.last_name}" in labels.text
            assert f"{donor.address}" in labels.text
            assert f"{donor.postal_code} {donor.city}" in labels.text

    @pytest.mark.parametrize("medal_id", (-1, -33, 99, 16))
    def test_envelope_invalid_medal(self, user, testapp, medal_id):
        medal = db.session.get(Medals, 1)
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        page.forms["printEnvelopeForm"].fields["medal_id"][0].value = medal_id
        labels = page.forms["printEnvelopeForm"].submit().follow()

        assert "Odesl치na nevalidn칤 data." in labels.text


class TestAwardPrepExport:
    @pytest.mark.parametrize("medal_id", range(1, 8))
    def test_award_prep_xlsx_table(self, user, testapp, medal_id):
        medal = db.session.get(Medals, medal_id)
        donation_centers = DonationCenter.query.all()
        dc_names = [dc.title for dc in donation_centers]
        login(user, testapp)
        page = testapp.get(url_for("donor.award_prep", medal_slug=medal.slug))
        table_data = page.click(description="St치hnout tabulku pro odb캩rn치 m칤sta").body
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(table_data)
            workbook = load_workbook(tmp.name)

        for sheetname in workbook.sheetnames:
            assert "roztridit" == sheetname or sheetname in dc_names
            assert workbook[sheetname]["A1"].value == "Jm칠no"
            assert workbook[sheetname]["B1"].value == "P콏칤jmen칤"

            for x in range(2, 11):
                first_name = workbook[sheetname][f"A{x}"].value
                if first_name:
                    assert first_name in page
                else:
                    break
                last_name = workbook[sheetname][f"B{x}"].value
                assert last_name in page
